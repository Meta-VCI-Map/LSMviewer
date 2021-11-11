# risk_scores_calculation/forms.py

from django.shortcuts import render,redirect
from django.http import JsonResponse
from rest_framework.viewsets import ViewSet
from rest_framework.parsers import JSONParser
from .forms import FileForm
import os
import time
from openpyxl import Workbook, load_workbook
from pathlib import Path
import pandas as pd
import xlsxwriter
import nibabel as nib
from scipy.ndimage import labeled_comprehension
from numpy import sum, mean, array, unique, nonzero, clip, log10
import shutil

try:
    '''Development'''
    from LSMviewer.settings import BASE_DIR, MEDIA_ROOT
    from risk_scores_calculation.serializers import CombinedSerializer
except:
    '''Deployment'''
    from LSMviewer.LSMviewer.deployment_settings import BASE_DIR, STATIC_ROOT, MEDIA
    from LSMviewer.risk_scores_calculation.serializers import CombinedSerializer

# Create your views here.
def filefield_upload(request,  *args, **kwargs):
    '''Image upload by users'''
    global img_obj

    '''delete the files that are kept on the server for more than 1 hour'''
    hours = 1
    secs = hours * 60 * 60
    secs_aday = 24 * 60 * 60
    days = secs / secs_aday
    seconds = time.time() - (days * secs_aday)

    try:
        '''Deployment'''
        path_dir = STATIC_ROOT
    except:
        '''Development'''
        path_dir = os.getcwd()
    '''excel files'''
    for file in os.listdir(path_dir):
        if file.endswith('.xlsx'):
            file_path = os.path.join(path_dir, file)
            ctime = os.stat(file_path).st_ctime
            if seconds >= ctime:
                delete_file_from_server(file_path)

    try:
        '''Deployment'''
        path_dir = STATIC_ROOT
    except:
        '''Development'''
        path_dir = f"{MEDIA_ROOT}"
    '''nifti files'''
    for file in os.listdir(path_dir):
        if ("MNI152_T1_1mm_brain_uint8" not in file) and ("location_impact_score_atlas" not in file) and ("network_impact_score_combined_atlas" not in file):
            if file.endswith('.nii.gz') or file.endswith('.nii') or file.startswith('Error'):
                file_path = os.path.join(path_dir, file)
                ctime = os.stat(file_path).st_ctime
                if seconds >= ctime:
                    delete_file_from_server(file_path)

    '''save file uploaded by user'''
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            img_obj = form.instance
            img_obj.save()
            try:
                '''Deployment'''
                source = os.path.join(img_obj.image.name)
                destination = os.path.join(STATIC_ROOT, img_obj.image.name)
                shutil.copy(source, destination)
            except:
                pass
            return render(request, 'risk_score_calculation.html', {'form': form, 'img_obj': img_obj, 'filename': img_obj.image.name})
    else:
        form = FileForm()
    return render(request, 'risk_score_calculation.html', {'form': form})


def delete_file_from_server(path):
    os.remove(os.path.join(path))


class RequestResultViewSet(ViewSet):
    """
    API endpoint
    """
    def calculate_location_score(request):
        '''Location Impact Score'''

        try:
            '''Development'''
            media_dir = MEDIA_ROOT
            atlas_dir = os.path.join(BASE_DIR, "static/")
        except:
            '''Deployment'''
            media_dir = os.getcwd()
            atlas_dir = STATIC_ROOT

        if request.is_ajax and request.method == "POST":
            print("Request:", request)

            '''file uploaded by user'''
            request_data = JSONParser().parse(request)
            infarct_obj = CombinedSerializer(request_data).data['infarct_image']
            session_name = img_obj.image.name
            for file in os.listdir(f"{media_dir}"):
                if file == infarct_obj['filename']:
                    session_name = file
            print("session/file: ", session_name)

            try:
                '''load file data'''
                uncompressed_filepath = f"{media_dir}/{session_name}"
                filepath = uncompressed_filepath + '.gz'
                shutil.copy(uncompressed_filepath, filepath)
                try:
                    '''try loading the compressed data'''
                    infarct_file = nib.load(filepath)
                except:
                    '''if an exception is raised, load the uncompressed data'''
                    infarct_file = nib.load(uncompressed_filepath)
            except:
                return JsonResponse({"LocationImpactScore": -500}, status=200)

            infarct_data = infarct_file.get_fdata()
            infarct_info = infarct_file.header
            infarct_shape = infarct_info['dim']
            infarct_dimsNum = infarct_shape[0]
            infarct_xDim = infarct_shape[1]
            infarct_yDim = infarct_shape[2]
            infarct_zDim = infarct_shape[3]
            infarct_xVoxel = infarct_shape[5]
            infarct_yVoxel = infarct_shape[6]
            infarct_zVoxel = infarct_shape[7]

            '''load atlas data'''
            coeff_file = f"{atlas_dir}/{'location_impact_score_atlas.nii.gz'}"
            coeff_data = (nib.load(coeff_file)).get_fdata()
            coeff_info = (nib.load(coeff_file)).header
            coeff_shape = coeff_info['dim']
            coeff_dimsNum = coeff_shape[0]
            coeff_xDim = coeff_shape[1]
            coeff_yDim = coeff_shape[2]
            coeff_zDim = coeff_shape[3]
            coeff_xVoxel = coeff_shape[5]
            coeff_yVoxel = coeff_shape[6]
            coeff_zVoxel = coeff_shape[7]

            try:
                if ~((infarct_data==0) | (infarct_data==1)).all() :
                    print("Infarct map is not binary")
                    return JsonResponse({"LocationImpactScore": -400}, status=200)
                if infarct_dimsNum != coeff_dimsNum:
                    print("Infarct map is not of the same dimensions as the atlas")
                    return JsonResponse({"LocationImpactScore": -100}, status=200)
                if infarct_xDim != coeff_xDim or infarct_yDim != coeff_yDim or infarct_zDim != coeff_zDim:
                    print("Infarct dimensions is not of the same shape as the atlas")
                    return JsonResponse({"LocationImpactScore": -200}, status=200)
                if infarct_xVoxel != coeff_xVoxel or infarct_yVoxel != coeff_yVoxel or infarct_zVoxel != coeff_zVoxel:
                    print("Infarct map has not the same voxels dimension as the atlas")
                    return JsonResponse({"LocationImpactScore": -300}, status=200)

                '''calculate the Location Impact Score'''
                location_impact_score = labeled_comprehension(
                                            input = coeff_data,
                                            labels = infarct_data,
                                            index = None,
                                            func = mean, out_dtype = float, default = 0.)
                print("Risk score =", location_impact_score)
                return JsonResponse({"LocationImpactScore": "%.4f" % location_impact_score}, status=200)

            except:
                return JsonResponse({"LocationImpactScore": 0}, status=400)
        return JsonResponse({"LocationImpactScore": 0}, status=400)


    def calculate_network_score(request):
        '''Network Impact Score'''

        try:
            '''Development'''
            media_dir = MEDIA_ROOT
            atlas_dir = os.path.join(BASE_DIR, "static/")
        except:
            '''Deployment'''
            media_dir = os.getcwd()
            atlas_dir = STATIC_ROOT
        '''read the .xlsx files for the region volumes and the hub scores'''
        hub = load_workbook(f"{atlas_dir}/{'hubscore.xlsx'}")
        AALvolumes = load_workbook(f"{atlas_dir}/{'newAALvolumes.xlsx'}")
        sheet_hub = hub.active
        sheet_volumes = AALvolumes.active

        '''get the length of the file'''
        row_count = sheet_hub.max_row
        column_count = sheet_hub.max_column

        '''save region names and hub scores to lists'''
        volume_name_list = []
        hubscore_list = []
        AALvolumes_list = []
        for i in range(1, row_count + 1):
            for j in range(1, column_count + 1):
                if i == 1 and j <= 90:
                    volume_name_list.append(sheet_hub.cell(row=i, column=j).value)
                elif i == 2 and j <= 90:
                    hubscore_list.append(sheet_hub.cell(row=i, column=j).value)
                    AALvolumes_list.append(sheet_volumes.cell(row=i, column=j).value)

        if request.is_ajax and request.method == "POST":
            print("Request:", request)

            '''file uploaded by user'''
            request_data = JSONParser().parse(request)
            infarct_obj = CombinedSerializer(request_data).data['infarct_image']
            session_name = img_obj.image.name
            for file in os.listdir(f"{media_dir}"):
                if file == infarct_obj['filename']:
                    session_name = file
            print("session/file: ", session_name)

            try:
                '''load file data'''
                uncompressed_filepath = f"{media_dir}/{session_name}"
                filepath = uncompressed_filepath + '.gz'
                shutil.copy(uncompressed_filepath, filepath)
                try:
                    '''try loading the compressed data'''
                    infarct_file = nib.load(filepath)
                except:
                    '''if an exception is raised, load the uncompressed data'''
                    infarct_file = nib.load(uncompressed_filepath)
            except:
                return JsonResponse({"NetworkImpactScore": -500}, status=200)

            infarct_data = infarct_file.get_fdata()
            infarct_info = infarct_file.header
            infarct_shape = infarct_info['dim']
            infarct_dimsNum = infarct_shape[0]
            infarct_xDim = infarct_shape[1]
            infarct_yDim = infarct_shape[2]
            infarct_zDim = infarct_shape[3]
            infarct_xVoxel = infarct_shape[5]
            infarct_yVoxel = infarct_shape[6]
            infarct_zVoxel = infarct_shape[7]

            '''load atlas data'''
            region_file = f"{atlas_dir}/{'network_impact_score_combined_atlas.nii.gz'}"
            region_data = (nib.load(region_file)).get_fdata()
            region_info = (nib.load(region_file)).header
            region_shape = region_info['dim']
            region_dimsNum = region_shape[0]
            region_xDim = region_shape[1]
            region_yDim = region_shape[2]
            region_zDim = region_shape[3]
            region_xVoxel = region_shape[5]
            region_yVoxel = region_shape[6]
            region_zVoxel = region_shape[7]

            try:
                if ~((infarct_data==0) | (infarct_data==1)).all() :
                    print("Infarct map is not binary")
                    return JsonResponse({"NetworkImpactScore": -400}, status=200)
                if infarct_dimsNum != region_dimsNum:
                    print("Infarct map is not of the same dimensions as the atlas")
                    return JsonResponse({"NetworkImpactScore": -100}, status=200)
                if infarct_xDim != region_xDim or infarct_yDim != region_yDim or infarct_zDim != region_zDim:
                    print("Infarct dimensions is not of the same shape as the atlas")
                    return JsonResponse({"NetworkImpactScore": -200}, status=200)
                if infarct_xVoxel != region_xVoxel or infarct_yVoxel != region_yVoxel or infarct_zVoxel != region_zVoxel:
                    print("Infarct map has not the same voxels dimensions as the atlas")
                    return JsonResponse({"NetworkImpactScore": -300}, status=200)


                '''Proceed with the Network impact score'''
                volumesNum = len(AALvolumes_list)
                impact_score = []

                labels = unique(region_data)
                '''drop the label of the background'''
                labels = labels[labels != 0]
                '''calculate the number of voxels of each region'''
                score_infarct = labeled_comprehension(
                                            infarct_data,
                                            region_data,
                                            labels,
                                            sum, float, 0.)
                score_infarct_volume = score_infarct / 1000.

                '''compute the risk score of each region'''
                each_volume = 1
                while each_volume < volumesNum + 1 :
                    region_volume = AALvolumes_list[ each_volume - 1 ]
                    infarct_fraction = score_infarct_volume[ each_volume - 1 ] / region_volume
                    '''clip fraction to [anything, 1.]'''
                    infarct_fraction = clip(infarct_fraction, None, 1.)
                    risk_score = infarct_fraction * hubscore_list[ each_volume - 1 ]
                    impact_score.append(risk_score)
                    each_volume += 1

                '''the network impact score is the maximum among the risk scores'''
                network_impact_score = max(impact_score)
                print(network_impact_score)

                '''check if the network impact score is zero.
                 If yes, the log10 is undefined'''
                if network_impact_score == 0.0:
                    return JsonResponse({"NetworkImpactScore": 0}, status=200)

                '''apply the log10'''
                log_network_impact_score = log10( network_impact_score )
                print("score: ",network_impact_score,"log: ", log_network_impact_score)

                nonzero_regions = nonzero(impact_score)
                nonzero_regions_list = list(nonzero_regions[0])
                corrected_list = nonzero_regions_list

                '''save the IDs of the regions that have an impact score other than zero
                    and make a file
                '''
                i = 0
                while i < len(corrected_list):
                    corrected_list[i] = corrected_list[i] +1
                    i += 1

                '''find the region (ID and name) of the maximum impact score'''
                max_index = list(impact_score).index(network_impact_score) + 1
                max_index_region_name = volume_name_list[max_index - 1]
                print(max_index, max_index_region_name)

                '''write data to .xlsx'''
                if session_name.endswith('.nii.gz'):
                    filename_nii = Path(session_name).stem
                    filename = Path(filename_nii).stem
                elif session_name.endswith('.nii'):
                    filename = Path(session_name).stem
                results = filename + "_network_impact_score_info.xlsx"
                print(results)
                saved_data_to_xlsx = xlsxwriter.Workbook(results)
                worksheet = saved_data_to_xlsx.add_worksheet("score")
                worksheet.write(0, 0, img_obj.image.name)
                worksheet.write(0, 1, "Impact Score")

                i = 0
                while i < len(corrected_list):
                    worksheet.write(i + 1, 0, volume_name_list[ corrected_list[i] -1 ] )
                    worksheet.write(i + 1, 1, impact_score[ corrected_list[i] -1 ])
                    i += 1
                worksheet.write(i+1, 0, "Max Impact Score")
                worksheet.write(i+1, 1, network_impact_score)
                worksheet.write(i+2, 0, "Region Name")
                worksheet.write(i+2, 1, max_index_region_name)
                worksheet.write(i+3, 0, "Log10")
                worksheet.write(i+3, 1, log_network_impact_score)
                saved_data_to_xlsx.close()

                return JsonResponse({"NetworkImpactScore": "%.4f" % network_impact_score,
                                     "Log": "%.4f" % log_network_impact_score,
                                     "MaxIndexRegionName": max_index_region_name,
                                     "ExcelFilename": results},
                                    status=200)
            except:
                return JsonResponse({"NetworkImpactScore": 0}, status=400)

        return JsonResponse({"NetworkImpactScore": 0}, status=400)


    def download_network_score(request):
        '''Download the excel file with the network score detailed results'''

        if request.is_ajax and request.method == "POST":

            request_data = JSONParser().parse(request)
            results = request_data['ExcelFilename']

            info_file = pd.read_excel(results)
            print(info_file)
            df_info_file = info_file.to_json()

            return JsonResponse({"InfoFile": df_info_file,
                                "Filename": info_file.columns[0],
                                "ExcelFilename": results},
                                status=200)