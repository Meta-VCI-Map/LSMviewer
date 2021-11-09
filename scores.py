# -*- coding: utf-8 -*-
import os
import numpy as np
import nibabel as nib
from scipy.ndimage import labeled_comprehension
from openpyxl import load_workbook
from LSMviewer import *

def location_score(infarct_data, filename):
    loc_img = os.path.join(os.getcwd(),'location_impact_score_atlas.nii.gz')
    loc_coefficient_image = nib.load(loc_img)
    loc_coefficient_data = loc_coefficient_image.get_fdata()
    loc_img_shape = loc_coefficient_data.shape 
    inf_img_shape = infarct_data.shape
    
    if loc_img_shape != inf_img_shape:            
        Ui_MainWindow.warning_message(len(globals.getFilenamesList), globals.nis, globals.lis, shape=False, fileName=filename)
        return None
    else:        
        location_impact_score = labeled_comprehension(
                                            input = loc_coefficient_data,
                                            labels = infarct_data,
                                            index = None,
                                            func = np.mean, out_dtype = float, default = 0.)          
        return "%.4f" % location_impact_score
    
    
def network_score(infarct_data, filename, nis_workbook):
    regionNum = 90
                
    '''Fisrt, read the xlsx files'''
    hub = load_workbook(os.path.join(os.getcwd(),'hubscore.xlsx'))
    AALvolumes = load_workbook(os.path.join(os.getcwd(),'newAALvolumes.xlsx'))
    sheet_hub = hub.active
    sheet_volumes = AALvolumes.active

    '''get the length of the file'''
    row_count = sheet_hub.max_row
    column_count = sheet_hub.max_column

    '''save volume names and hubscores to lists'''
    volume_name_list = []
    hubscore_list = []
    AALvolumes_list = []
    for i in range(1, row_count + 1):
        for j in range(1, column_count + 1):
            if i == 1 and j <= regionNum:
                volume_name_list.append(sheet_hub.cell(row=i, column=j).value)
            elif i == 2 and j <= regionNum:
                hubscore_list.append(sheet_hub.cell(row=i, column=j).value)
                AALvolumes_list.append(sheet_volumes.cell(row=i, column=j).value)
    
    net_img = os.path.join(os.getcwd(),'network_impact_score_combined_atlas.nii.gz')
    net_coefficient_image = nib.load(net_img)
    net_coefficient_data = net_coefficient_image.get_fdata()
    
    labels = np.unique(net_coefficient_data)
    '''drop the label of the background'''
    labels = labels[labels != 0]
    impact_score = []
    
    net_img_shape = net_coefficient_data.shape
    inf_img_shape = infarct_data.shape
    
    if net_img_shape != inf_img_shape:
        Ui_MainWindow.warning_message(len(globals.getFilenamesList), globals.nis, globals.lis, shape=False, fileName=filename)
        return None, None
    else:
        score_infarct = labeled_comprehension(
                                            input = infarct_data,
                                            labels = net_coefficient_data,
                                            index = labels,
                                            func = np.sum, out_dtype = float, default = 0.)
        score_infarct_volume = score_infarct / 1000.
        
        each_volume = 1
        while each_volume < regionNum + 1 :
            region_volume = AALvolumes_list[ each_volume - 1 ]
            infarct_fraction = score_infarct_volume[ each_volume - 1 ] / region_volume
            '''clip fraction to [anything, 1.]'''
            infarct_fraction = np.clip(infarct_fraction, None, 1.)
            risk_score = infarct_fraction * hubscore_list[ each_volume - 1 ]
            impact_score.append(risk_score)
            each_volume += 1
        network_impact_score = max(impact_score)
        
        '''check if the max impact score is zero.
         If yes, the log10 is undefined'''
        if network_impact_score == 0.0:
            undefined_score = "Undefined"
            undefined_name = "Undefined"
            return undefined_score, undefined_name

        '''apply the log10'''
        log_network_impact_score = np.log10( network_impact_score )
        print("score: ",network_impact_score,"log: ", log_network_impact_score)
        
        nonzero_regions = np.nonzero(impact_score)
        nonzero_regions_list = list(nonzero_regions[0])
        corrected_list = nonzero_regions_list

        '''save the IDs of the regions that have an impact score other than zero
            and make a file'''
        i = 0
        while i < len(corrected_list):
            corrected_list[i] = corrected_list[i] +1
            i += 1

        '''find the region (ID and name) of the maximum impact score'''
        max_index = list(impact_score).index(network_impact_score) + 1
        max_index_region_name = volume_name_list[max_index - 1]

        Ui_MainWindow.write_nis_to_file(nis_workbook, filename, corrected_list, volume_name_list, impact_score, 
                           network_impact_score, max_index_region_name, log_network_impact_score)

        return "%.4f" % log_network_impact_score, max_index_region_name